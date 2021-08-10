from NaverNewsCrawler import NaverNewsCrawler
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
import smtplib
import re
import json

##개인정보 json파일 열기
with open('my.json') as f: 
    config = json.load(f)

SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465

##json파일에서 참조
SMTP_USER = config['email']
SMTP_PASSWORD = config['password']

##메일 보내는 함수(default)
def send_mail(name, addr, subject, contents, attachment=None):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()


## 검색할 키워드를 사용자에게 입력 받기
keyword = input('뉴스를 검색할 키워드를 입력하세요. : ') 
crawler = NaverNewsCrawler(keyword)

## 저장 파일명 엑셀파일인지 체크 후 키워드 관련 뉴스 리스트 파일형태로 저장
while True:
    news_list_file = input('뉴스를 저장할 파일 이름을 입력하세요(엑셀) : ')
    if (news_list_file.find('.')== -1):
        news_list_file = news_list_file + '.xlsx'
    if (news_list_file[len(news_list_file)-5:] == '.xlsx'):
        crawler.get_news(news_list_file)
        break
    print('올바르지 않은 형식입니다. 다시 입력해주세요')

contents = "'" + keyword + "' 관련 뉴스 크롤링 리스트 파일 입니다."

## 이메일리스트 불러오기
wb = load_workbook('email_list.xlsx',read_only=True)
data = wb.active

## 이메일리스트의 이름, 이메일 주소로 메일 보내기
first_row = True  # 첫번째 행 판별 기본값
for row in data.iter_rows():
    if first_row:
        first_row = False
        continue
    name = row[1].value
    addr = row[2].value 
    send_mail(name,addr,keyword+' 관련 뉴스 리스트입니다.',contents,news_list_file)